
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import pytest
import openpyxl

from Pages.HomePage import HomePage
from Pages.LoginPage import Login

import logging
from Utilities.LogUtil import Logger

# object = className(to know from which test,log level)
log = Logger(__name__,logging.INFO)


def get_data():

   workbook = openpyxl.load_workbook("../Excel/TestData.xlsx")
   sheet = workbook["Sheet1"]
   totalRows = sheet.max_row
   totalColm = sheet.max_column
   mainList = [] # initialize empty list

   for i in range(2, totalRows+1): # Start with row 2 because 1 contains header
       dataList = [] # to keep each set of data
       for j in range(1, totalColm+1):
           data = sheet.cell(row=i, column=j).value
           dataList.insert(j,data) # Adding 1st coloumn data
       mainList.insert(i, dataList)
   print("Printing all the values from mainlist",mainList)
   return mainList


class Test_Login: # Test "T" should be in uppercase for class
    @pytest.mark.usefixtures("log_on_failure")
    @pytest.mark.parametrize("username,password",get_data())
    def test_login(self,username, password):

        log.logger.info("Test Login started")
        loginPage = Login(self.driver) # calling LoginPage class Login
        loginPage.enterLoginDetails(username, password)
        log.logger.info("Test Login successfully executed")




